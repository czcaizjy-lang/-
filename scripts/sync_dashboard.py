#!/usr/bin/env python3
"""
业绩看板数据提取脚本
直接读取上游 Excel 文件，生成 dashboard_data.json

数据口径：
  - 达人花名册（昵称/抖音号/机构）来自「星辞业绩」和「星辞自达号业绩」
  - 所有数值指标（GMV/消耗/退款等）从「X月直播数据」日流水实时聚合
  - 汇总/达人/机构均按当月过滤；趋势图取近 30 天滚动窗口
"""

import json
import os
import subprocess as sp
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 直接读取上游文件
XLSX_PATH = '/Users/xiaocao/Desktop/蕉下文件/业绩追击/by月业绩/6月业绩/6月业绩追击（纯直播）.xlsx'
OUTPUT_PATH = '/Users/xiaocao/CC/每日业绩自动统计/data/dashboard_data.json'
MAIN_AGENCIES = ['自达号', '集米文化', '紫语', '花开满路', '太古', '亦初', '直属']
ZDH_SUB_AGENCIES = ['花开自达号', '集米自达号', '太古自达号', '九三自达号', '直属自达号']
TREND_DAYS = 30  # 趋势图近 N 天


def find_live_sheet(wb):
    """自动查找日流水 sheet（名称包含「直播数据」）"""
    for name in wb.sheetnames:
        if '直播数据' in name:
            return name
    raise ValueError(f'未找到直播数据 sheet，可用 sheets: {wb.sheetnames}')


def safe_float(v):
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def run():
    wb = load_workbook(XLSX_PATH, data_only=True)

    # ═══════════════════════════════════════════
    # === 1. 达人花名册（仅读昵称/抖音号/机构）===
    # ═══════════════════════════════════════════

    # 1a. 全部达人花名册（星辞业绩）
    ws_anchor = wb['星辞业绩']
    all_roster = {}  # douyin_id → {name, agency}
    all_roster_order = []  # 保持原始顺序
    for r in range(3, ws_anchor.max_row + 1):
        douyin_id = ws_anchor.cell(r, 2).value
        if not douyin_id:
            continue
        did = str(douyin_id)
        all_roster[did] = {
            '主播昵称': str(ws_anchor.cell(r, 1).value or ''),
            '机构': str(ws_anchor.cell(r, 9).value or '其他'),
        }
        all_roster_order.append(did)

    # 1b. 自达号花名册（星辞自达号业绩）
    ws_zdh = wb['星辞自达号业绩']
    zdh_roster = {}  # douyin_id → {name, agency}
    zdh_roster_order = []
    zdh_id_to_sub = {}  # 保留兼容
    for r in range(3, ws_zdh.max_row + 1):
        douyin_id = ws_zdh.cell(r, 2).value
        if not douyin_id:
            continue
        did = str(douyin_id)
        sub_org = str(ws_zdh.cell(r, 10).value or '其他')
        zdh_roster[did] = {
            '主播昵称': str(ws_zdh.cell(r, 1).value or ''),
            '机构': sub_org,
        }
        zdh_roster_order.append(did)
        zdh_id_to_sub[did] = sub_org

    # ═══════════════════════════════════════════
    # === 2. 读取日流水（X月直播数据）===
    # ═══════════════════════════════════════════
    live_sheet_name = find_live_sheet(wb)
    ws_live = wb[live_sheet_name]
    print(f'  日流水 sheet: {live_sheet_name}')

    # 按抖音号 + 日期维度累加
    daily_gmv = defaultdict(lambda: defaultdict(float))       # douyin_id → date → gmv
    daily_paid = defaultdict(lambda: defaultdict(float))      # douyin_id → date → paid
    daily_refund = defaultdict(lambda: defaultdict(float))    # douyin_id → date → refund
    daily_ad = defaultdict(lambda: defaultdict(float))        # douyin_id → date → ad_cost
    daily_commission = defaultdict(lambda: defaultdict(float)) # douyin_id → date → commission
    daily_duration = defaultdict(lambda: defaultdict(float))   # douyin_id → date → duration(min)
    live_name_map = {}  # 抖音号 → 昵称（日流水覆盖最全）

    all_dates_set = set()

    for r in range(2, ws_live.max_row + 1):
        douyin_id_raw = ws_live.cell(r, 3).value  # C列
        nickname_raw = ws_live.cell(r, 2).value    # B列
        dt_val = ws_live.cell(r, 4).value           # D列

        if not douyin_id_raw or not dt_val:
            continue

        douyin_id = str(douyin_id_raw)
        date_key = str(dt_val)[:10].replace('/', '-')
        all_dates_set.add(date_key)

        # 昵称映射
        if nickname_raw:
            live_name_map[douyin_id] = str(nickname_raw)

        gmv = safe_float(ws_live.cell(r, 26).value)           # Z: 直播间成交金额
        paid = safe_float(ws_live.cell(r, 27).value)           # AA: 支付金额
        refund = safe_float(ws_live.cell(r, 32).value)         # AF: 退款金额
        ad_bind = safe_float(ws_live.cell(r, 44).value)        # AR: 投放消耗(绑定)
        ad_beitou = safe_float(ws_live.cell(r, 45).value)      # AS: 投放消耗(被投)
        commission = safe_float(ws_live.cell(r, 34).value)     # AH: 预估佣金支出
        duration = safe_float(ws_live.cell(r, 6).value)        # F: 直播时长(分钟)

        # 消耗：被投优先，0 回退绑定（现逻辑）
        ad_cost = ad_beitou if ad_beitou > 0 else ad_bind

        daily_gmv[douyin_id][date_key] += gmv
        daily_paid[douyin_id][date_key] += paid
        daily_refund[douyin_id][date_key] += refund
        daily_ad[douyin_id][date_key] += ad_cost if gmv > 0 else 0
        daily_commission[douyin_id][date_key] += commission
        daily_duration[douyin_id][date_key] += duration

    all_dates = sorted(all_dates_set)

    # ═══════════════════════════════════════════
    # === 3. 确定当前月份 ===
    # ═══════════════════════════════════════════
    latest_date = all_dates[-1]
    current_month = latest_date[:7]  # e.g. "2026-07"
    current_month_dates = [d for d in all_dates if d.startswith(current_month)]
    print(f'  当前月份: {current_month} | 最新日期: {latest_date} | 当月天数: {len(current_month_dates)}')

    # ═══════════════════════════════════════════
    # === 4. 达人月度指标（当月聚合）===
    # ═══════════════════════════════════════════

    def sum_month(daily_dict, douyin_id):
        """汇总某达人当月所有日期的值"""
        return round(sum(
            daily_dict.get(douyin_id, {}).get(d, 0)
            for d in current_month_dates
        ), 2)

    def count_active_days(daily_gmv_dict, douyin_id):
        """统计当月 GMV > 0 的天数"""
        return sum(1 for d in current_month_dates
                   if daily_gmv_dict.get(douyin_id, {}).get(d, 0) > 0)

    # 构建全部达人列表（当月有产出的）
    all_anchor_monthly = {}
    for douyin_id in all_roster:
        gmv_val = sum_month(daily_gmv, douyin_id)
        paid_val = sum_month(daily_paid, douyin_id)
        refund_val = sum_month(daily_refund, douyin_id)
        ad_val = sum_month(daily_ad, douyin_id)
        commission_val = sum_month(daily_commission, douyin_id)
        duration_total = sum_month(daily_duration, douyin_id)
        active_days = count_active_days(daily_gmv, douyin_id)
        settle_val = round(paid_val - refund_val, 2)  # 结算 = 支付 - 退款

        info = all_roster[douyin_id]
        all_anchor_monthly[douyin_id] = {
            '主播昵称': info['主播昵称'],
            '主播抖音号': douyin_id,
            '机构': info['机构'],
            '开播天数': active_days,
            '日均开播时长（小时）': round(duration_total / 60 / active_days, 1) if active_days > 0 else 0,
            '直播GMV': gmv_val,
            '直播支付GMV': paid_val,
            '直播退款GMV': refund_val,
            '直播结算GMV': settle_val,
            '结算率': round(settle_val / gmv_val, 4) if gmv_val > 0 else 0,
            'ROI': round(gmv_val / ad_val, 2) if ad_val > 0 else 0,
            '佣金支出': commission_val,
            '投放消耗金额': ad_val,
        }

    # 构建自达号达人列表
    zdh_anchor_monthly = {}
    for douyin_id in zdh_roster:
        gmv_val = sum_month(daily_gmv, douyin_id)
        paid_val = sum_month(daily_paid, douyin_id)
        refund_val = sum_month(daily_refund, douyin_id)
        ad_val = sum_month(daily_ad, douyin_id)
        commission_val = sum_month(daily_commission, douyin_id)
        duration_total = sum_month(daily_duration, douyin_id)
        active_days = count_active_days(daily_gmv, douyin_id)
        settle_val = round(paid_val - refund_val, 2)

        info = zdh_roster[douyin_id]
        zdh_anchor_monthly[douyin_id] = {
            '主播昵称': info['主播昵称'],
            '主播抖音号': douyin_id,
            '机构': info['机构'],
            '开播天数': active_days,
            '日均开播时长（小时）': round(duration_total / 60 / active_days, 1) if active_days > 0 else 0,
            '直播GMV': gmv_val,
            '直播支付GMV': paid_val,
            '直播退款GMV': refund_val,
            '直播结算GMV': settle_val,
            '结算率': round(settle_val / gmv_val, 4) if gmv_val > 0 else 0,
            'ROI': round(gmv_val / ad_val, 2) if ad_val > 0 else 0,
            '佣金支出': commission_val,
            '投放消耗金额': ad_val,
        }

    # ═══════════════════════════════════════════
    # === 5. 汇总卡片（当月全量）===
    # ═══════════════════════════════════════════
    total_gmv = sum(a['直播GMV'] for a in all_anchor_monthly.values())
    total_paid = sum(sum_month(daily_paid, did) for did in all_roster)
    total_refund = sum(a['直播退款GMV'] for a in all_anchor_monthly.values())
    total_settle = round(total_paid - total_refund, 2)
    total_ad = sum(a['投放消耗金额'] for a in all_anchor_monthly.values())
    total_commission = sum(a['佣金支出'] for a in all_anchor_monthly.values())

    summary = {
        '主播昵称': '汇总',
        '直播GMV': round(total_gmv, 2),
        '直播支付GMV': round(total_paid, 2),
        '直播退款GMV': round(total_refund, 2),
        '直播结算GMV': total_settle,
        '结算率': round(total_settle / total_gmv, 4) if total_gmv > 0 else 0,
        '佣金支出': round(total_commission, 2),
        '投放消耗金额': round(total_ad, 2),
    }

    # 自达号汇总
    zdh_total_gmv = sum(a['直播GMV'] for a in zdh_anchor_monthly.values())
    zdh_total_paid = sum(sum_month(daily_paid, did) for did in zdh_roster)
    zdh_total_refund = sum(a['直播退款GMV'] for a in zdh_anchor_monthly.values())
    zdh_total_settle = round(zdh_total_paid - zdh_total_refund, 2)
    zdh_total_ad = sum(a['投放消耗金额'] for a in zdh_anchor_monthly.values())

    zdh_summary = {
        '直播GMV': round(zdh_total_gmv, 2),
        '直播支付GMV': round(zdh_total_paid, 2),
        '直播退款GMV': round(zdh_total_refund, 2),
        '直播结算GMV': zdh_total_settle,
        '结算率': round(zdh_total_settle / zdh_total_gmv, 4) if zdh_total_gmv > 0 else 0,
        '消耗金额': round(zdh_total_ad, 2),
        'ROI': round(zdh_total_gmv / zdh_total_ad, 2) if zdh_total_ad > 0 else 0,
    }

    # ═══════════════════════════════════════════
    # === 6. 机构汇总（当月，按机构分组）===
    # ═══════════════════════════════════════════
    agency_data = defaultdict(lambda: {
        'anchors': set(),
        'gmv': 0.0, 'refund': 0.0, 'paid': 0.0,
        'ad': 0.0, 'commission': 0.0,
    })

    for douyin_id, a in all_anchor_monthly.items():
        agency = a['机构']
        agency_data[agency]['anchors'].add(douyin_id)
        agency_data[agency]['gmv'] += a['直播GMV']
        agency_data[agency]['refund'] += a['直播退款GMV']
        agency_data[agency]['paid'] += sum_month(daily_paid, douyin_id)
        agency_data[agency]['ad'] += a['投放消耗金额']
        agency_data[agency]['commission'] += a['佣金支出']

    agencies = []
    for agency in MAIN_AGENCIES:
        if agency not in agency_data:
            continue
        d = agency_data[agency]
        n = len(d['anchors'])
        settle = round(d['paid'] - d['refund'], 2)
        agencies.append({
            '机构': agency,
            '机构达人数': n,
            '直播GMV': round(d['gmv'], 2),
            '人均直播GMV': round(d['gmv'] / n, 2) if n > 0 else 0,
            '直播退款GMV': round(d['refund'], 2),
            '直播结算GMV': settle,
            '结算率': round(settle / d['gmv'], 4) if d['gmv'] > 0 else 0,
            '佣金支出': round(d['commission'], 2),
            '投放消耗金额': round(d['ad'], 2),
        })

    # 子机构汇总（自达号）
    zdh_sub_data = defaultdict(lambda: {
        'anchors': set(),
        'gmv': 0.0, 'refund': 0.0, 'paid': 0.0, 'ad': 0.0,
    })
    for douyin_id, a in zdh_anchor_monthly.items():
        sub = a['机构']
        zdh_sub_data[sub]['anchors'].add(douyin_id)
        zdh_sub_data[sub]['gmv'] += a['直播GMV']
        zdh_sub_data[sub]['refund'] += a['直播退款GMV']
        zdh_sub_data[sub]['paid'] += sum_month(daily_paid, douyin_id)
        zdh_sub_data[sub]['ad'] += a['投放消耗金额']

    zdh_sub_agencies = []
    for sub in ZDH_SUB_AGENCIES:
        if sub not in zdh_sub_data:
            continue
        d = zdh_sub_data[sub]
        n = len(d['anchors'])
        settle = round(d['paid'] - d['refund'], 2)
        zdh_sub_agencies.append({
            '机构': sub,
            '机构达人数': n,
            '直播GMV': round(d['gmv'], 2),
            '人均直播GMV': round(d['gmv'] / n, 2) if n > 0 else 0,
            '直播退款GMV': round(d['refund'], 2),
            '直播结算GMV': settle,
            '投放消耗金额': round(d['ad'], 2),
            'ROI': round(d['gmv'] / d['ad'], 2) if d['ad'] > 0 else 0,
        })

    # ═══════════════════════════════════════════
    # === 7. 趋势数据（近 30 天）===
    # ═══════════════════════════════════════════
    trend_dates_full = all_dates[-TREND_DAYS:] if len(all_dates) >= TREND_DAYS else all_dates
    trend_dates = [f"{int(d[5:7])}/{int(d[8:10])}" for d in trend_dates_full]
    date_map = {trend_dates_full[i]: trend_dates[i] for i in range(len(trend_dates_full))}

    # 每日全量聚合
    total_gmv_daily = defaultdict(float)
    total_paid_daily = defaultdict(float)
    total_refund_daily = defaultdict(float)
    daily_by_agency = defaultdict(lambda: defaultdict(float))

    # 自达号每日
    zdh_gmv_daily = defaultdict(float)
    zdh_paid_daily = defaultdict(float)
    zdh_refund_daily = defaultdict(float)
    zdh_daily_by_sub = defaultdict(lambda: defaultdict(float))

    for d in trend_dates_full:
        for douyin_id, by_date in daily_gmv.items():
            gmv = by_date.get(d, 0)
            paid = daily_paid[douyin_id].get(d, 0)
            refund = daily_refund[douyin_id].get(d, 0)
            total_gmv_daily[d] += gmv
            total_paid_daily[d] += paid
            total_refund_daily[d] += refund

            # 机构归属（全部）
            agency = all_roster.get(douyin_id, {}).get('机构', '其他')
            daily_by_agency[agency][d] += gmv

            # 自达号
            if douyin_id in zdh_roster:
                zdh_gmv_daily[d] += gmv
                zdh_paid_daily[d] += paid
                zdh_refund_daily[d] += refund
                sub = zdh_roster[douyin_id]['机构']
                zdh_daily_by_sub[sub][d] += gmv

    def daily_list(daily_dict):
        """将 defaultdict 转为按 trend_dates_full 顺序的列表（单位：万）"""
        return [round(daily_dict.get(d, 0) / 10000, 2) for d in trend_dates_full]

    # ═══════════════════════════════════════════
    # === 8. 自达号子机构每日 ROI ===
    # ═══════════════════════════════════════════
    zdh_anchor_ids_by_sub = defaultdict(list)
    for douyin_id, info in zdh_roster.items():
        zdh_anchor_ids_by_sub[info['机构']].append(douyin_id)

    zdh_daily_roi_by_sub = {}
    for sub in ZDH_SUB_AGENCIES:
        aids = zdh_anchor_ids_by_sub.get(sub, [])
        roi_vals = []
        for d in trend_dates_full:
            gmv_sum = sum(daily_gmv.get(aid, {}).get(d, 0) for aid in aids)
            ad_sum = sum(daily_ad.get(aid, {}).get(d, 0) for aid in aids)
            roi_vals.append(round(gmv_sum / ad_sum, 2) if ad_sum > 0 else None)
        zdh_daily_roi_by_sub[sub] = roi_vals

    # 整体 ROI
    all_zdh_ids = list(zdh_roster.keys())
    zdh_daily_roi_overall = []
    for d in trend_dates_full:
        gmv_sum = sum(daily_gmv.get(aid, {}).get(d, 0) for aid in all_zdh_ids)
        ad_sum = sum(daily_ad.get(aid, {}).get(d, 0) for aid in all_zdh_ids)
        zdh_daily_roi_overall.append(round(gmv_sum / ad_sum, 2) if ad_sum > 0 else None)

    # ═══════════════════════════════════════════
    # === 9. 人员分天（久酒 / 雅宁 / 星辞 + 其他）===
    # ═══════════════════════════════════════════
    person_sheet_names = ['久酒业绩', '雅宁业绩', '星辞业绩']
    person_douyin_sets = {}  # person_name → set of douyin_ids
    person_name_map = {}     # douyin_id → 昵称（从人员 sheet A 列）

    for sheet_name in person_sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        douyin_ids = set()
        for r in range(2, ws.max_row + 1):
            did = ws.cell(r, 2).value  # B列
            if did:
                did_str = str(did)
                douyin_ids.add(did_str)
                name = ws.cell(r, 1).value  # A列
                if name:
                    person_name_map[did_str] = str(name)
        if douyin_ids:
            person_name = sheet_name.replace('业绩', '')
            person_douyin_sets[person_name] = douyin_ids

    # 「其他」= 全部日流水达人 - 三人集合
    three_person_ids = set()
    for ids in person_douyin_sets.values():
        three_person_ids |= ids
    other_ids = set(daily_gmv.keys()) - three_person_ids

    person_daily = {'dates': trend_dates}
    person_daily_gmv = {'dates': trend_dates}
    person_daily_refund = {'dates': trend_dates}

    for person_name, douyin_ids in person_douyin_sets.items():
        paid_vals, gmv_vals, refund_vals = [], [], []
        for d in trend_dates_full:
            paid_vals.append(round(sum(daily_paid.get(did, {}).get(d, 0) for did in douyin_ids) / 10000, 2))
            gmv_vals.append(round(sum(daily_gmv.get(did, {}).get(d, 0) for did in douyin_ids) / 10000, 2))
            refund_vals.append(round(sum(daily_refund.get(did, {}).get(d, 0) for did in douyin_ids) / 10000, 2))
        person_daily[person_name] = paid_vals
        person_daily_gmv[person_name] = gmv_vals
        person_daily_refund[person_name] = refund_vals

    # 「其他」
    if other_ids:
        paid_vals, gmv_vals, refund_vals = [], [], []
        for d in trend_dates_full:
            paid_vals.append(round(sum(daily_paid.get(did, {}).get(d, 0) for did in other_ids) / 10000, 2))
            gmv_vals.append(round(sum(daily_gmv.get(did, {}).get(d, 0) for did in other_ids) / 10000, 2))
            refund_vals.append(round(sum(daily_refund.get(did, {}).get(d, 0) for did in other_ids) / 10000, 2))
        person_daily['其他'] = paid_vals
        person_daily_gmv['其他'] = gmv_vals
        person_daily_refund['其他'] = refund_vals

    # ═══════════════════════════════════════════
    # === 10. 下探数据 ===
    # ═══════════════════════════════════════════

    # 10a. agency_top5_anchors
    agency_top5_anchors = {}
    for agency in MAIN_AGENCIES:
        # 该机构下所有达人按当月 GMV 排序
        agency_anchors = [
            (did, all_anchor_monthly[did]['直播GMV'])
            for did in all_roster
            if all_anchor_monthly[did]['机构'] == agency
        ]
        agency_anchors.sort(key=lambda x: x[1], reverse=True)
        top5 = []
        for douyin_id, _ in agency_anchors[:5]:
            name = all_roster[douyin_id]['主播昵称']
            vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
            top5.append({'name': name, 'douyin_id': douyin_id, 'daily_paid': vals})
        if top5:
            agency_top5_anchors[agency] = top5

    # 10b. zdh_top5_by_sub
    zdh_top5_by_sub = {}
    for sub in ZDH_SUB_AGENCIES:
        sub_anchors = [
            (did, zdh_anchor_monthly[did]['直播GMV'])
            for did in zdh_roster
            if zdh_anchor_monthly[did]['机构'] == sub
        ]
        sub_anchors.sort(key=lambda x: x[1], reverse=True)
        top5 = []
        for douyin_id, _ in sub_anchors[:5]:
            name = zdh_roster[douyin_id]['主播昵称']
            vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
            top5.append({'name': str(name), 'douyin_id': douyin_id, 'daily_paid': vals})
        if top5:
            zdh_top5_by_sub[sub] = top5

    # 10c. person_anchor_detail（人员下探）
    person_anchor_detail = {}
    for person_name, douyin_ids in person_douyin_sets.items():
        anchor_list = []
        for douyin_id in douyin_ids:
            info = all_roster.get(douyin_id, {})
            name = person_name_map.get(douyin_id) or info.get('主播昵称') or douyin_id
            vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
            anchor_list.append({'name': str(name), 'douyin_id': douyin_id, 'daily_paid': vals})
        anchor_list.sort(key=lambda x: sum(x['daily_paid']), reverse=True)
        person_anchor_detail[person_name] = anchor_list

    # 10d. all_anchor_daily（全部达人下探）
    all_anchor_daily = []
    for douyin_id in daily_gmv:
        info = all_roster.get(douyin_id, {})
        name = live_name_map.get(douyin_id) or info.get('主播昵称') or douyin_id
        vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
        all_anchor_daily.append({'name': str(name), 'douyin_id': douyin_id, 'daily_paid': vals})
    all_anchor_daily.sort(key=lambda x: sum(x['daily_paid']), reverse=True)

    # 10e. zdh_anchor_detail（自达号下探）
    zdh_anchor_detail = []
    for douyin_id in zdh_roster:
        info = zdh_roster[douyin_id]
        vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
        zdh_anchor_detail.append({
            'name': str(info['主播昵称'] or douyin_id),
            'douyin_id': douyin_id,
            'daily_paid': vals
        })
    zdh_anchor_detail.sort(key=lambda x: sum(x['daily_paid']), reverse=True)

    # anchor_daily_paid（按抖音号，用于前端局部展示）
    anchor_daily_paid_out = {}
    for douyin_id in daily_paid:
        anchor_daily_paid_out[douyin_id] = {
            date_map[d]: round(daily_paid[douyin_id].get(d, 0) / 10000, 2)
            for d in trend_dates_full
        }

    # anchor_daily_roi
    anchor_daily_roi = {}
    for douyin_id in daily_gmv:
        anchor_daily_roi[douyin_id] = {}
        for d in trend_dates_full:
            g = daily_gmv[douyin_id].get(d, 0)
            a = daily_ad.get(douyin_id, {}).get(d, 0)
            anchor_daily_roi[douyin_id][date_map[d]] = round(g / a, 2) if a > 0 else 0

    # ═══════════════════════════════════════════
    # === 11. 构建最终 JSON ===
    # ═══════════════════════════════════════════

    # 排序：全部达人按当月 GMV 降序
    anchors_sorted = sorted(all_anchor_monthly.values(), key=lambda x: x['直播GMV'], reverse=True)
    zdh_anchors_sorted = sorted(zdh_anchor_monthly.values(), key=lambda x: x['直播GMV'], reverse=True)

    dashboard_data = {
        'summary': summary,
        'anchors': anchors_sorted,
        'agencies': agencies,
        'trend': {
            'dates': trend_dates,
            'daily_total_gmv': daily_list(total_gmv_daily),
            'daily_total_paid': daily_list(total_paid_daily),
            'daily_total_refund': daily_list(total_refund_daily),
            'agencies': list(daily_by_agency.keys()),
            'daily_by_agency': {
                ag: daily_list(daily_by_agency[ag])
                for ag in daily_by_agency
            },
            'agency_totals': round(sum(
                sum(daily_by_agency[ag].values()) for ag in daily_by_agency
            ) / 10000, 2),
            'pie_data': [
                {'name': ag, 'value': round(sum(daily_by_agency[ag].values()) / 10000, 2)}
                for ag in daily_by_agency
            ],
            'person_daily': person_daily,
            'person_daily_gmv': person_daily_gmv,
            'person_daily_refund': person_daily_refund,
            'anchor_daily_paid': anchor_daily_paid_out,
            'anchor_daily_roi': anchor_daily_roi,
            'person_anchor_detail': person_anchor_detail,
            'all_anchor_daily': all_anchor_daily,
            'agency_top5_anchors': agency_top5_anchors,
        },
        'zidahao': {
            'summary': zdh_summary,
            'sub_agencies': zdh_sub_agencies,
            'anchors': zdh_anchors_sorted,
            'daily_gmv': daily_list(zdh_gmv_daily),
            'daily_paid': daily_list(zdh_paid_daily),
            'daily_refund': daily_list(zdh_refund_daily),
            'daily_by_sub': {
                sub: daily_list(zdh_daily_by_sub[sub])
                for sub in ZDH_SUB_AGENCIES
            },
            'pie_data': [
                {
                    'name': sub,
                    'value': round(sum(zdh_daily_by_sub[sub].values()) / 10000, 2)
                }
                for sub in ZDH_SUB_AGENCIES
            ],
            'top5_by_sub': zdh_top5_by_sub,
            'daily_roi_by_sub': zdh_daily_roi_by_sub,
            'daily_roi_overall': zdh_daily_roi_overall,
            'anchor_detail': zdh_anchor_detail,
        },
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    print(f'✓ 数据已更新到 {OUTPUT_PATH}')
    print(f'  - 当前月份: {current_month}')
    print(f'  - 汇总 GMV: ¥{summary["直播GMV"]:,.2f}')
    print(f'  - 结算 GMV: ¥{summary["直播结算GMV"]:,.2f}')
    print(f'  - 达人数量: {len(anchors_sorted)}')
    print(f'  - 机构数量: {len(agencies)}')
    print(f'  - 趋势窗口: {trend_dates[0]} ~ {trend_dates[-1]} ({len(trend_dates)} 天)')
    print(f'  - 自达号 GMV: ¥{zdh_summary["直播GMV"]:,.2f} ({len(zdh_anchors_sorted)} 达人, {len(zdh_sub_agencies)} 子机构)')
    print(f'  - 人员分天: {list(person_douyin_sets.keys())} + 其他({len(other_ids)}达人)')

    # 自动构建独立看板页面
    script_dir = os.path.dirname(os.path.abspath(__file__))
    build_script = os.path.join(script_dir, 'build_standalone.py')
    if os.path.exists(build_script):
        print('')
        sp.run(['python3', build_script], cwd=script_dir)

if __name__ == '__main__':
    run()
